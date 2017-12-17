#! /usr/bin/env python
# -*- coding: utf-8 -*-

"""
@Author:Lich
@Time:  2017/12/5 11:07
@Description: 将excel件转存到txt文件中
"""
from export_data import build_features_lib
from openpyxl import load_workbook
from file_path_constant import *
import codecs
import re
import time
import os
import sys
reload(sys)
sys.setdefaultencoding('utf8')

__author__ = 'Lich'


def import_data_from_excel():
    print u'正在导入，请等待...'
    start_time = time.time()
    log_info = {}
    for dir_name in dirs:
        path = os.path.join(mac_test_path, dir_name)
        wb = load_workbook(os.path.join(path, dir_name + r'.xlsx'))
        print wb.sheetnames
        sheet = wb.get_sheet_by_name("sheet1")

        tmp_path = os.path.join(mac_path, dir_name)
        a = 0
        for row in sheet['A']:
            try:
                f = codecs.open(os.path.join(tmp_path, str(a) + r'.txt'), 'w', 'utf-8', errors='ignore')
                txt = str(row.value).decode('ISO-8859-15').encode('utf-8')
                txt = re.sub(r'[^\x00-\x7F]+', '', txt)  # 去除所有非ASCII字符
                if len(txt) < 100 :
                    print '---------------------None'
                    continue
                a += 1
                f.write(txt)
            except IOError, e:
                print dir_name, u'从Excel导入数据异常，IOError ', e.message
            except UnicodeDecodeError, e:
                print dir_name, u'从Excel导入数据异常，UnicodeDecodeError ', e.message
            except UnicodeEncodeError, e:
                print dir_name, u'从Excel导入数据异常，UnicodeEncodeError ', e.message
            finally:
                f.close()
        log_info[dir_name] = a
        end_time = time.time()
    for item in log_info:
        print item, u'类新闻共%d篇' % log_info[item]
    print u'excel文件导入TXT完成！ 一共耗时%.4f秒' % (end_time - start_time)


'''
从feature目录下的特征库读取特征
'''


def import_features_from_lib():
    features = []
    all_features_words = set([])
    if not os.path.exists(mac_f_path):
        # build_features_lib()
        print 'import_features_from_lib ----------------!!!'
        return 0
    else:
        for dir_name in dirs:
            try:
                f = codecs.open(os.path.join(mac_f_path, dir_name + '.txt'), 'rb')
                txt = f.read().decode('ISO-8859-15').encode('utf-8')
                txt = re.sub(r'[^\x00-\x7F]+', '', txt)  # 去除所有非ASCII字符
                lst = txt.split(' ')
                print dir_name, "特征包含共%d个词"%len(lst)
                all_features_words = all_features_words | set(lst)
                features.append((lst, dir_name))  # [(lst1,cat1),(lst2,cat2),...,(lst7,cat7)]
            except IOError, e:
                print dir_name, u'特征库读取异常，IOError ', e.message
            except UnicodeDecodeError, e:
                print dir_name, u'特征库读取异常，UnicodeDecodeError ', e.message
            except UnicodeEncodeError, e:
                print dir_name, u'特征库读取异常，UnicodeEncodeError ', e.message
            finally:
                f.close()
    return features, all_features_words
